#------------------------------------------------------------------------------------------------------#
#  ___                  _ __  __            _                       
# / _ \ _ __ ___  _ __ (_)  \/  | ___ _ __ | |_ ___  _ __
#| | | | '_ ` _ \| '_ \| | |\/| |/ _ \ '_ \| __/ _ \| '__|
#| |_| | | | | | | | | | | |  | |  __/ | | | || (_) | |
# \___/|_| |_| |_|_| |_|_|_|  |_|\___|_| |_|\__\___/|_|
#
# SMS/LMS with command line parser and tcp socket protocol support
#------------------------------------------------------------------------------------------------------
import warnings
warnings.filterwarnings("ignore")
import pandas as pd
import pymysql
import pymysql.cursors
import json
import datetime
import sys
import requests
from openpyxl import load_workbook
import vmsvclib, vmbotlib
from vmsvclib import *

global edx_api_header,edx_api_url,max_iu

piece = lambda txtstr,seperator,pos : txtstr.split(seperator)[pos]
string2date = lambda x,y : datetime.datetime.strptime(x,y).date()
str2date = lambda x : string2date(x,"%d/%m/%Y")
dmy_str = lambda v : piece(v,'-',2) + '/' + piece(v,'-',1) + '/' + piece(v,'-',0)
ymd_str = lambda v : piece(v,'-',0) + '/' + piece(v,'-',1) + '/' + piece(v,'-',2)
getcohort = lambda x : '-'.join(x.split(':')[1].replace('+','-').split('-')[:-1][1:])
getmodulecode = lambda x : '-'.join(x.split(':')[1].replace('+','-').split('-')[:-1][1:][:-1])

def course_header(client_name,course_id):
    global edx_api_header,edx_api_url
    if course_id=="":
        return ["","",""]
    qry = f"select * from playbooks where client_name = '{client_name}' and course_id='{course_id}' limit 1;"
    df = rds_df(qry)
    if (df is not None) and (len(df)>0):
        df1 = df.T
        df1.columns = ['col']
        records = [x for x in df1.col]
        pillar = records[2]
        course_code = records[3]
        module_code = records[4]
        return [pillar,course_code,module_code]
    else:
        module_code = getmodulecode(course_id)
        qry = f"select * from course_module where client_name = '{client_name}' and module_code='{module_code}';"
        df = rds_df(qry)
        if (df is not None) and (len(df)==1):
            df.columns = vmsvclib.get_columns("course_module")
            pillar = [x for x in df.pillar][0]
            course_code = [x for x in df.course_code][0]
            return [pillar,course_code,module_code]
    url = f"{edx_api_url}/course/fetch/info/{course_id}"
    result = dict()
    response = requests.get(url,headers=edx_api_header)
    data = eval(response.content.decode('utf-8') if response.status_code==200 else "[{}]")
    result = data[0] if len(data)>0 else dict()
    output = ([result['pillar'],result['courseCode'],result['moduleCode']] if len(result)>=4 else ['','','']) if isinstance(result,dict) else ['','','']
    return output

def edx_coursename(course_id):
    global edx_api_header,edx_api_url
    coursename = ""
    url = f"{edx_api_url}/course/fetch/name"
    response = requests.post(url,data=course_id,headers=edx_api_header,verify=False)
    if response.status_code==200:
        data = response.content.decode('utf-8')
        if len(data)>0:
            coursename = str(data)
    return coursename

def edx_day0(course_id):
    global edx_api_header,edx_api_url
    start_date = ""
    url = f"{edx_api_url}/course/fetch/startdate"
    response = requests.post(url,data=course_id,headers=edx_api_header,verify=False)
    if response.status_code==200:
        data = response.content.decode('utf-8')
        if len(data)>0:
            start_date = datetime.datetime.strptime(str(data)[:10],"%Y-%m-%d").date()
    return start_date

def edx_courseid(cohort_id):
    global edx_api_header,edx_api_url
    course_id = ""
    url = f"{edx_api_url}/course/fetch/id"
    response = requests.post(url,data=cohort_id,headers=edx_api_header,verify=False)
    if response.status_code==200:
        data = response.content.decode('utf-8')
        if len(data)>0:
            course_id = str(data)
    return course_id

def edx_mcqcnt(course_id,client_name):
    qry = "SELECT bb.mcq_cnt as cnt from playbooks aa INNER JOIN course_module bb "
    qry += "ON aa.pillar=bb.pillar AND aa.course_code=bb.course_code AND aa.module_code=bb.module_code "
    qry += f"AND aa.client_name=bb.client_name WHERE aa.client_name='{client_name}' AND aa.course_id = '{course_id}';"
    cnt = rds_param(qry)
    cnt = 0 if str(cnt)=='' else cnt
    return cnt

def edx_ascnt(course_id,client_name):
    qry = "SELECT bb.as_cnt as cnt from playbooks aa INNER JOIN course_module bb "
    qry += "ON aa.pillar=bb.pillar AND aa.course_code=bb.course_code AND aa.module_code=bb.module_code "
    qry += f"AND aa.client_name=bb.client_name WHERE aa.client_name='{client_name}' AND aa.course_id = '{course_id}';"
    cnt = rds_param(qry)
    cnt = 0 if str(cnt)=='' else cnt
    return cnt

def edx_userdata(course_id):
    global edx_api_header,edx_api_url
    df = pd.DataFrame.from_dict({'student_id':[],'course_id':[],'username':[],'email':[]})
    url = f"{edx_api_url}/course/fetch/users"
    syslog(f"calling api {url} via requests.post")
    response = requests.post(url,data=course_id,headers=edx_api_header,verify=False)
    syslog(f"completed with status code {response.status_code}")
    if response.status_code==200:
        data = response.content.decode('utf-8')
        if len(data) > 0:
            userinfo = json.loads(data)
            rows_cnt = len(list(userinfo))
            user_list = []
            course_list = []
            username_list = []
            email_list = []
            if rows_cnt > 0:
                for usr in userinfo:
                    if course_id in [x['course_id'] for x in usr['enrolments']]:
                        user_list.append(usr['user_id'])
                        course_list.append(course_id)
                        username_list.append(usr['username'])
                        email_list.append(usr['email'])
            data = {'student_id':user_list,'course_id':course_list,'username':username_list,'email':email_list}
            df = pd.DataFrame.from_dict(data)
    return df

def student_course_list(student_id):
    global edx_api_header,edx_api_url
    df = pd.DataFrame.from_dict({'course_id':[]})
    url = f"{edx_api_url}/user/fetch/{student_id}"
    response = requests.get(url,headers=edx_api_header)
    if response.status_code==200:
        data = response.content.decode('utf-8')
        if len(data)>0:
            userinfo = json.loads(response.content.decode('utf-8'))
            course_list = [x['course_id'] for x in userinfo['enrolments']]
            df = pd.DataFrame.from_dict({'course_id':course_list})
    return df

def edx_grade(course_id,student_id=0):
    global edx_api_header,edx_api_url
    df = pd.DataFrame.from_dict({'student_id':[],'grade':[]})
    if student_id==0:
        url = f"{edx_api_url}/user/fetch/grades/list"
    else:
        url = f"{edx_api_url}/user/fetch/grades/list/{student_id}"
    syslog(f"calling api {url} via requests.post")
    response = requests.post(url,data=course_id,headers=edx_api_header,verify=False)
    syslog(f"completed with status code {response.status_code}")
    if response.status_code==200:
        data = response.content.decode('utf-8')
        if len(data)==0:
            data = []
            syslog(f"{url} returns no data")
        else:
            data = eval(data)
        stud = [x['student_id'] for x in data]
        grade = [x['grade'] for x in data]
        data = {'student_id' : stud,'grade' : grade }
        df = pd.DataFrame.from_dict(data)
    return df

def edx_mcqinfo(client_name,course_id,student_id=0):
    global edx_api_header,edx_api_url
    getnumstr = lambda q : int('0'+(( ''.join([x for x in q if x.isnumeric() or x==' '])).strip()).split(' ')[0])

    df_mcq = pd.DataFrame.from_dict( {'client_name':[],'course_id':[],'student_id':[],'score':[], \
                                      'mcq':[],'qn':[],'attempts':[],'avgscore': []} )
    if student_id==0:
        url = f"{edx_api_url}/user/fetch/mcq/scores/list"
    else:
        url = f"{edx_api_url}/user/fetch/mcq/scores/list/{student_id}"
    syslog(f"calling api {url} via requests.post course_id = {course_id}")
    response = requests.post(url,data=course_id,headers=edx_api_header,verify=False)
    syslog(f"completed with status code {response.status_code}")
    if response.status_code==200:
        data = response.content.decode('utf-8')
        if str(data) == "":
            data = []
            syslog(f"{url} returns no data")
        else:
            data = json.loads(data)
        course_id_list = []
        student_id_list = []
        iu_list = []
        qn_list = []
        grade_list = []
        att_list = []
        avgscore_list = []
        for rec in [ x for x in list(data) ]:
            qn = 99
            pp = 0
            qtitle = ""
            if 'student_id' in list(rec):
                sid = int(rec['student_id'])
            else:
                sid = 0
            if 'score' in list(rec):
                sc = rec['score']
                sc = 0 if sc is None else sc
            else:
                sc = 0
            if 'points_possible' in list(rec):
                pp = rec['points_possible']
                pp = 0 if pp is None else pp
            if 'options_display_name' in list(rec):
                qtitle = rec['options_display_name']
            if 'option_display_name' in list(rec):
                qtitle = rec['option_display_name']
            if qtitle.startswith('Q'):
                qn = getnumstr(qtitle)
            if 'IU' in list(rec):
                r = rec['IU']
                if (r is None) or (r==""):
                    syslog(f"IU = None on {course_id} student_id = {sid}")
                    if 'chapter_title' in list(rec):
                        iu = getnumstr(rec['chapter_title'])
                    else:
                        iu = 0
                else:
                    iu = int(rec['IU'])
            else:
                iu = 0
            state = eval(rec['state'].replace('null','""').replace('false','0').replace('true','1'))
            qnscore = 0
            if 'correct_map' in list(state):
                map_list = list(state['correct_map'])
                if len(map_list)==0:
                    syslog('correct_map segment was empty')
                else:
                    statekey = map_list[0]
                    correctness = state['correct_map'][statekey]['correctness']
                    qnscore = 1.0 if correctness=='correct' else 0
            grade = 1.0 if pp==0 else float(sc/pp)
            attempts = 0 if qnscore == 0 else ( state['attempts'] if 'attempts' in list(state) else 0)
            course_id_list.append(rec['course_id'])
            student_id_list.append(rec['student_id'])
            iu_list.append(iu)
            qn_list.append(qn)
            grade_list.append(qnscore)
            att_list.append(int(attempts))
            avgscore_list.append(grade)
        client_list = [client_name for x in iu_list]
        data = {'client_name':client_list,'course_id': course_id_list,'student_id':student_id_list, \
                'score':grade_list ,'mcq':iu_list ,'qn':qn_list ,'attempts': att_list,'avgscore': avgscore_list}
        df_mcq = pd.DataFrame.from_dict(data)
    return df_mcq

def edx_assignment_score(course_id,student_id=0):
    global edx_api_header,edx_api_url
    get_iu = lambda x : (x+' ').split(' ')[0].replace('IU','').replace(':','').strip()
    if student_id==0:
        url = f"{edx_api_url}/user/fetch/assignment/scores/list"
    else:
        url = f"{edx_api_url}/user/fetch/assignment/scores/list/{student_id}"
    syslog(f"calling api {url} via requests.post")
    response = requests.post(url,data=course_id,headers=edx_api_header,verify=False)
    syslog(f"completed with status code {response.status_code}")
    df = pd.DataFrame.from_dict({'student_id':[],'score':[],'points_possible':[],'IU': [],'attempts': []})
    if response.status_code==200:
        data = response.content.decode('utf-8')
        data = data.replace('"IU":null','"IU":"0"')
        rec = eval(data) if len(data)>0 else []
        if rec == []:
            syslog(f"{url} returns no data")
        else:
            student_id_list = [x['student_id'] for x in rec]
            grade_list = [x['score'] for x in rec]
            pp_list = [x['points_possible'] for x in rec]
            iu_list = [ get_iu(x['IU']) for x in rec]
            att_list = [ 0 if x['score']==0 else x['attempts'] for x in rec]
            data = {'student_id': student_id_list,'score': grade_list,'points_possible':pp_list,'IU': iu_list,'attempts': att_list}
            df = pd.DataFrame.from_dict(data)
            df['is_num'] = df.apply(lambda x: 1 if x['IU'].isnumeric() else 0,axis=1)
            df.drop(df[ df.is_num == 0 ].index,inplace=True)
            df.drop(columns=['is_num'],inplace=True)
    return df

def sms_df(course_id,student_id):
    global edx_api_header,edx_api_url
    if student_id==0:
        url = f"{edx_api_url}/user/attendance/fetch/course/id"
    else:
        url = f"{edx_api_url}/user/attendance/fetch/{student_id}"
    syslog(f"calling api {url} via requests.post")
    response = requests.post(url,data=course_id,headers=edx_api_header,verify=False)
    syslog(f"completed with status code {response.status_code}")
    if response.status_code==200:
        data = response.content.decode('utf-8')
        if len(data)>0:
            if student_id==0:
                att_dict=eval(str(data))
                att_list = []
                for email in att_dict:
                    e_list = list(eval(str(att_dict[email])))
                    for att in e_list:
                        att['email'] = email
                    att_list += e_list
                df = pd.DataFrame.from_records(att_list)
                return df
            df = pd.DataFrame.from_records(list(eval(str(data))))
            if list(df.columns) != []:
                return df
    return None

def sms_datelist(df):
    if df is None:
        return []
    df0 = df[df.attendance_type_desc.str.lower()=='attend']
    dlist = [x.split(' ')[0] for x in df0.timetable_date]
    try:
        date_list = [string2date(x,"%m/%d/%Y") for x in dlist]
    except:
        date_list = [string2date(x,"%d/%m/%Y") for x in dlist]
    return date_list

def sms_attendance(course_id,student_id):
    global edx_api_header,edx_api_url
    if student_id==0:
        return []
    url = f"{edx_api_url}/user/attendance/fetch/{student_id}"
    syslog(f"calling api {url} via requests.post")
    response = requests.post(url,data=course_id,headers=edx_api_header,verify=False)
    syslog(f"completed with status code {response.status_code}")
    date_list = []
    if response.status_code==200:
        data = response.content.decode('utf-8')
        if len(data)>0:
            result  = eval(str(data))
            try:
                date_list = [string2date(x['timetable_date'].split(' ')[0],"%m/%d/%Y") \
                             for x in result if x['attendance_type_desc']=='Attend']
            except:
                date_list = [string2date(x['timetable_date'].split(' ')[0],"%d/%m/%Y") \
                             for x in result if x['attendance_type_desc']=='Attend']
    else:
        syslog(f"responsed with no data! code = {response.status_code}")
    return date_list

def sms_missingdates(client_name,course_id,student_id,cols,date_list):
    f2flag = lambda x,y : ('' if y in date_list else x) if f2f_map[x]==1 else ''
    if (student_id==0) or (client_name==""):
        return []
    missed_fsf = []
    df = rds_df(f"select * from stages where courseid = '{course_id}' and client_name = '{client_name}';")
    if df is not None:
        df.columns = cols
        stgid_list = [x for x in df.id]
        fsf_dates = [string2date(x,"%d/%m/%Y") for x in df.startdate]
        stage_list = [x for x in df.stage]
        f2f_list = [x for x in df.f2f]
        f2f_map = dict(zip(stage_list,f2f_list))
        arr_stgf2f = dict(zip(stage_list,fsf_dates))
        missing_fsf = [f2flag(stage_list[n],fsf_dates[n]) for n in range(len(stgid_list))]
    return missing_fsf

def sms_att_rate(client_name,course_id,student_id,date_list):
    if (student_id==0) or (client_name==""):
        return 0
    query = f"select startdate from stages where client_name = '{client_name}' and courseid='{course_id}' "
    query += "and (f2f = 1) and STR_TO_DATE(startdate,'%d/%m/%Y') <= CURDATE();"
    cnt2 = 0
    df = rds_df(query)
    if df is None:
        return 0
    else:
        df.columns = ['startdate']
    dt_list = [string2date(x,"%d/%m/%Y") for x in df.startdate]
    cnt2 = len(dt_list)
    cnt1 = len([x for x in dt_list if x in date_list])
    att_rate = 0 if cnt2 == 0 else (cnt1/cnt2)
    return att_rate

def sms_pmstage(client_name,course_id):
    if (course_id==0) or (client_name==""):
        return []
    query = f"select count(*) as cnt from stages where client_name = '{client_name}' and courseid='{course_id}' "
    query += "and stage like 'PM%' and STR_TO_DATE(startdate,'%d/%m/%Y') <= CURDATE();"
    pmcounts = rds_param(query)
    pmstage = 0 if pmcounts == 0 else 1
    return pmstage

def search_course_list(keyword):
    global edx_api_header,edx_api_url
    course_list = []
    url = f"{edx_api_url}/course/search/course/list"
    syslog(f"calling api {url} via requests.post")
    response = requests.post(url,data=keyword,headers=edx_api_header,verify=False)
    syslog(f"completed with status code {response.status_code}")
    if response.status_code==200:
        data = response.content.decode('utf-8')
        if len(data) > 0:
            course_list = eval(data)
    return course_list

def edx_course_started(client_name,course_id):
    query = "SELECT count(*) as cnt FROM stages WHERE id=1 and client_name ="
    query += f"'{client_name}' AND courseid='{course_id}' AND "
    query += "STR_TO_DATE(stagedate,'%d/%m/%Y') <= CURDATE();"
    try:
        cnt = rds_param(query)
        cnt = int(cnt)
    except:
        cnt = 0
    return cnt

def edx_endofcourse(client_name,course_id):
    query = f"SELECT (case SUBSTRING(`name`,-3) when 'EOC' then 1 else 0 END) eoc FROM stages WHERE client_name ="
    query += f"'{client_name}' AND courseid='{course_id}' AND "
    query += "STR_TO_DATE(stagedate,'%d/%m/%Y') <= CURDATE() ORDER BY id DESC LIMIT 1;"
    try:
        result = rds_param(query)
        eoc = 1 if result=='' else int(result)
    except:
        eoc = 0
    return eoc

def eoc_date(client_name,course_id):
    query = f"SELECT stagedate FROM stages WHERE client_name ="
    query += f"'{client_name}' AND courseid='{course_id}' ORDER BY id DESC LIMIT 1;"
    result = ""
    try:
        result = rds_param(query)
        result = datetime.datetime.strptime(str(result),"%d/%m/%Y").date()
    except:
        pass
    return result

def edx_eocgap(client_name,course_id,gap=7):
    query = "select COUNT(*) as cnt from stages  where stage = 'EOC' and "
    query += f"STR_TO_DATE(stagedate,'%d/%m/%Y') BETWEEN (CURDATE() - INTERVAL {gap} DAY) AND CURDATE() "
    query += f" and client_name ='{client_name}' and courseid = '{course_id}';"
    try:
        results = rds_param(query)
        eocgap=int(results)
    except:
        eocgap = 0
    return eocgap

def edx_eocend(client_name,course_id,gap=7):
    query = "select COUNT(*) as cnt from stages  where stage = 'EOC' and "
    query += f"CURDATE() - STR_TO_DATE(stagedate,'%d/%m/%Y') = {gap} "
    query += f" and client_name ='{client_name}' and courseid = '{course_id}';"
    try:
        results = rds_param(query)
        eoc7=int(results)
    except:
        eoc7 = 0
    return eoc7

def update_schedule(course_id,client_name):
    dstr = lambda x : piece(piece(x.strip(),':',1),'+',2)
    dtype = lambda x : ('%d%b%Y' if len(dstr(x))==9 else '%d%B%Y') if dstr(x)[0].isdigit() else '%B%Y'
    [ pillar,course_code,module_code ] = course_header(client_name,course_id)
    if (pillar=="") or (course_code=="") or (module_code==""):
        syslog(f"The course setup for {course_id} is incomplete, update_schedule stopped")
        return False
    cohort_id = getcohort(course_id)
    stage_list = get_google_calendar(course_id,client_name)
    if stage_list == []:
        syslog("google calendar return nothing")
    try:
        stage_list = update_stage_table(stage_list,course_id,client_name,cohort_id,pillar,course_code,module_code)
        if stage_list == []:
            syslog(f"System calendar not match for {course_id}")
            return False
        dt = string2date(stage_list[0][3],"%d/%m/%Y")
        stg=[x for x in stage_list if x[0]=='EOC']
        eoc_date = string2date(stg[0][4],"%d/%m/%Y")
        date_today = datetime.datetime.now().date()
        days = (date_today - eoc_date).days
    except:
        syslog(f"update_stage_table failed for {course_id}")
        return False
    condqry = "client_name = '_c_' and courseid = '_x_';"
    condqry = condqry.replace('_c_',client_name)
    condqry = condqry.replace('_x_',course_id)
    qry = "select * from stages where " + condqry
    df = rds_df(qry)
    if df is None:
        stage_list = []
        days_list = []
        as_str = []
        mcq_str = []
    else:
        df.columns = get_columns("stages")
        stage_list = [x for x in df.name]
        days_list = [x for x in df.days]
        as_str = [x for x in df.assignment]
        mcq_str = [x for x in df.mcq]
    for n in range(len( days_list )):
        stg = stage_list[n]
        dd = days_list[n]
        stg_date = dt + datetime.timedelta( days = (dd-1) )
        stagedate = stg_date.strftime('%d/%m/%Y')
        qry = "update stages set days = _y_ where name = '_x_' and " + condqry
        qry = qry.replace('_x_',stg)
        query = qry.replace('_y_',str(dd))
        rds_update(query)
    for fld in ['mcq','assignment','IU']:
        query = f"UPDATE stages_master SET {fld}='0' WHERE client_name = '{client_name}' and {fld} = '';"
        rds_update(query)
    query = "update stages SET IU=IFNULL((SELECT b.IU FROM stages_master b WHERE b.client_name=client_name "
    query += f"AND b.module_code='{module_code}' AND b.stage=stage LIMIT 1) ,'0') "
    query += f"WHERE courseid = '{course_id}' AND client_name = '{client_name}';"
    rds_update(query)
    query = f"SELECT `name` FROM stages WHERE client_name = '{client_name}' AND courseid='{course_id}' AND "
    query += "STR_TO_DATE(startdate,'%d/%m/%Y') <= CURDATE() ORDER BY id DESC LIMIT 1;"
    current_stage = rds_param(query)
    current_stage = current_stage.replace("\r\n","")
    query = f"update userdata set stage = '{current_stage}' where " + condqry
    rds_update(query)
    query = f"update stages set stagedate = '' where (stagedate is null) and "
    query += condqry
    rds_update(query)
    query = f"update stages set startdate = '' where (startdate is null) and "
    query += condqry
    rds_update(query)
    for fld in ['mcq','assignment','IU']:
        query = f"UPDATE stages SET {fld}='0' WHERE client_name = '{client_name}' and courseid='{course_id}' and {fld} = '';"
        rds_update(query)
    query = "UPDATE userdata SET mcq_zero = '[]',mcq_failed = '[]',as_zero = '[]',as_failed = '[]',risk_level=0 "
    query += f" WHERE client_name='{client_name}' and courseid='{course_id}';"
    rds_update(query)
    qry = "SELECT max(SUBSTRING_INDEX(mcq, ',', -1)) AS max_mcq, "
    qry += "max(SUBSTRING_INDEX(assignment, ',', -1)) AS max_as "
    qry += f"FROM stages WHERE courseid = '{course_id}' AND client_name = '{client_name}';"
    df = rds_df(qry)
    pass_rate = rds_param("SELECT value from params where client_name = 'System' AND `KEY` = 'pass_rate';")
    pass_rate = float(pass_rate)
    if df is not None:
        df.columns = ['max_mcq','max_as']
        max_mcq = int(df.max_mcq.values[0])
        max_as = int(df.max_as.values[0])
        list_cond = [ f"(a.mcq_avg{n}>=0 AND a.mcq_avg{n} <{pass_rate} AND concat(s.mcq , ',') LIKE '%{n},%')" for n in range(1,max_mcq+1)]
        list_cond += [ f"(a.as_avg{n}>=0 AND a.as_avg{n} <{pass_rate} AND concat(s.assignment , ',') LIKE '%{n},%')" for n in range(1,max_as+1)]
        if len(list_cond)>0:
            mcq_cond = ' OR '.join(list_cond)
            qry = "UPDATE userdata a SET stage = IFNULL((SELECT s.name FROM stages s WHERE ( "
            qry += mcq_cond
            qry += f") AND a.courseid=s.courseid LIMIT 1),'{current_stage}') WHERE a.courseid = '{course_id}' AND a.client_name = '{client_name}';"
            rds_update(qry)
        qry = "UPDATE stages s INNER JOIN userdata a ON a.client_name=s.client_name AND a.courseid=s.courseid SET a.mcq_zero = Replace(CONCAT('[',"
        list_cond = [ f"if(a.mcq_avg{n}=0 AND CONCAT(s.mcq,',') LIKE '%{n},%','{n},','')," for n in range(1,max_mcq+1) ]
        if len(list_cond)>0:
            qry += ''.join(list_cond)
            qry += f"']'),',]',']') WHERE s.name = '{current_stage}' AND s.client_name = '{client_name}' AND s.courseid = '{course_id}';"
            rds_update(qry)
        qry = "UPDATE stages s INNER JOIN userdata a ON a.client_name=s.client_name AND a.courseid=s.courseid SET a.mcq_failed = Replace(CONCAT('[',"
        list_cond = [ f"if(a.mcq_avg{n}>0 AND a.mcq_avg{n}<{pass_rate} AND CONCAT(s.mcq,',') LIKE '%{n},%','{n},','')," for n in range(1,max_mcq+1) ]
        if len(list_cond)>0:
            qry += ''.join(list_cond)
            qry += f"']'),',]',']') WHERE s.name = '{current_stage}' AND s.client_name = '{client_name}' AND s.courseid = '{course_id}';"
            rds_update(qry)
        qry = "UPDATE stages s INNER JOIN userdata a ON a.client_name=s.client_name AND a.courseid=s.courseid SET a.as_zero = Replace(CONCAT('[',"
        list_cond = [ f"if(a.as_avg{n}=0 AND CONCAT(s.assignment,',') LIKE '%{n},%','{n},','')," for n in range(1,max_as+1) ]
        if len(list_cond)>0:
            qry += ''.join(list_cond)
            qry += f"']'),',]',']') WHERE s.name = '{current_stage}' AND s.client_name = '{client_name}' AND s.courseid = '{course_id}';"
            rds_update(qry)
        qry = "UPDATE stages s INNER JOIN userdata a ON a.client_name=s.client_name AND a.courseid=s.courseid SET a.as_failed = Replace(CONCAT('[',"
        list_cond = [ f"if(a.as_avg{n}>0 AND a.as_avg{n}<{pass_rate} AND CONCAT(s.assignment,',') LIKE '%{n},%','{n},','')," for n in range(1,max_as+1) ]
        if len(list_cond)>0:
            qry += ''.join(list_cond)
            qry += f"']'),',]',']') WHERE s.name = '{current_stage}' AND s.client_name = '{client_name}' AND s.courseid = '{course_id}';"
            rds_update(qry)
        qry = f"UPDATE userdata SET risk_level = 1 WHERE client_name = '{client_name}' AND courseid = '{course_id}' "
        qry += " AND NOT (mcq_zero='[]' AND mcq_failed='[]' AND as_zero='[]' AND as_failed='[]');"
        rds_update(qry)
    syslog(f"completed on {course_id}")
    return True

def update_assignment(course_id,client_name,student_id=0):
    condqry = f"client_name = '{client_name}' and courseid = '{course_id}';"
    as_cnt = edx_ascnt(course_id,client_name)
    if as_cnt<=0:
        syslog(f"edx_ascnt returns 0")
        return
    syslog(f"Found {as_cnt} IUs on {course_id}")
    df = edx_grade(course_id,student_id)
    if df is None:
        return False
    if len(df) == 0:
        stud_grade_dict = dict()
    else:
        stud_grade_dict = dict(zip([x for x in df.student_id],[x for x in df.grade]))
    df = edx_assignment_score(course_id,student_id)
    if len(df) == 0:
        sid_list = []
        score_list = []
        iu_list = []
        attempts_list = []
        pp_list = []
    else:
        if 'attempts' not in list(df.columns):
            df['attempts'] = 1
        sid_list = [x for x in df.student_id]
        score_list = [x for x in df.score]
        iu_list = [x for x in df.IU]
        attempts_list = [x for x in df.attempts]
        pp_list = [x for x in df.points_possible]
    stud_list = []
    iu_score = dict()
    iu_attempts = dict()
    for n in range(len(df)):
        sid = sid_list[n]
        iu_num = int(iu_list[n])
        pp_num = pp_list[n]
        score = score_list[n]
        attempts = attempts_list[n]
        if sid not in stud_list:
            stud_list.append(sid)
            iu_score[sid] = dict()
            iu_attempts[sid] = dict()
        iu_score[sid][iu_num] = 1 if pp_num == 0 else (int(score)/pp_num)
        attempts = 0 if score==0 else attempts
        iu_attempts[sid][iu_num] = attempts
    for sid in list(stud_grade_dict):
        if sid not in stud_list:
            iu_score[sid] = dict()
            iu_attempts[sid] = dict()
    stud_list = list(iu_score)
    for sid in stud_list:
        if sid not in list(stud_grade_dict):
            stud_grade_dict[sid] = 0
    for n in range(len(stud_list)):
        sid = stud_list[n]
        if sid not in list(stud_grade_dict):
            stud_grade_dict[sid] = 0
        updsqlavg = ''.join([ ", as_avg" + str(x) + " = " + str(iu_score[sid][x]) for x in list(iu_score[sid]) if x > 0])
        updsqlatt = ''.join([ ", as_attempts" + str(x) + " = " + str(iu_attempts[sid][x]) for x in list(iu_score[sid]) if x > 0 ])
        updsql = "update userdata set grade = " + str(stud_grade_dict[sid]) + updsqlavg + updsqlatt
        updsql += " where studentid = " + str(sid) + " and courseid = '" + course_id + "';"
        try:
            rds_update(updsql)
        except:
            errlog(f"SQL update error\n{updsql}")
    syslog("completed for course_id " + course_id)
    return True

def update_mcq(course_id,client_name,student_id=0):
    global max_iu
    cohort_id = getcohort(course_id)
    if student_id==0:
        condqry = f"client_name = '{client_name}' and courseid = '{course_id}'"
        cond_qry = f"client_name = '{client_name}' and course_id = '{course_id}'"
    else:
        condqry = f"client_name = '{client_name}' and courseid = '{course_id}' and studentid  = {student_id}"
        cond_qry = f"client_name = '{client_name}' and course_id = '{course_id}' and student_id  = {student_id}"
    syslog("edx_mcqcnt : " + course_id)
    mcqcnt = edx_mcqcnt(course_id,client_name)
    if mcqcnt<=0:
        syslog(f"edx_mcqcnt returns 0")
        return
    syslog(f"Found {mcqcnt} IUs on {course_id}")
    mcq_df = edx_mcqinfo(client_name,course_id,student_id)
    if mcq_df is None:
        syslog(f"edx_mcqinfo return none for {course_id}")
        return
    if len(mcq_df) > 0:
        query = "delete from mcq_data where " + cond_qry
        rds_update(query)
    else:
        pass
    df = mcq_df[['client_name','course_id','student_id','score','mcq','qn','attempts']]
    syslog(f"copydbtbl on mcq_data")
    copydbtbl(df,"mcq_data")
    query = "select client_name,course_id,student_id,mcq,count(*) as max,sum(score) as score,max(attempts) as max_attempts from mcq_data where "
    query += cond_qry
    query += " group by client_name,course_id,student_id,mcq"
    query += " order by client_name,course_id,student_id,mcq;"
    scoredf = rds_df(query)
    cnt = 0
    if scoredf is None :
        syslog(f"there is no data for table mcq_data")
    else:
        cnt = len(scoredf)
        if cnt > 0:
            query = "delete from mcq_score where " + condqry
            rds_update(query)
            scoredf.columns = ['client_name','courseid','studentid','mcq','max','score','max_attempts']
            syslog("copydbtbl : mcq_score")
            copydbtbl(scoredf,"mcq_score")
    if cnt == 0:
        max_mcq = 0
    else:
        query = "select max(mcq) AS maxmcq from mcq_score where " + condqry
        df = rds_df(query)
        max_mcq = 0
        if df is not None:
            df.columns = ['maxmcq']
            max_mcq = df['maxmcq'][0]
        if max_mcq > max_iu:
            max_mcq = 0

    if max_mcq==0:
        syslog("max_mcq = 0")
        return

    syslog("reset mcq attempts")
    updqry = "update userdata set "
    updqry += ','.join([ "mcq_attempts" + str(x) + " = 0"  for x in range( 1,max_mcq + 1 )]) + " where "
    updqry += condqry
    rds_update(updqry)

    syslog("update mcq attempts")
    query = "SELECT distinct student_id as sid FROM mcq_data WHERE " + cond_qry + " order by student_id;"
    df = rds_df(query)
    if df is None:
        sid_list = []
    else:
        df.columns = ['sid']
        sid_list = [x for x in df.sid]

    for sid in sid_list:
        query = f"SELECT mcq , max(attempts) as att FROM mcq_data WHERE client_name = '{client_name}' AND course_id = '{course_id}' AND student_id={sid} GROUP BY mcq"
        df = rds_df(query)
        if df is None:
            att_dict = {}
        else:
            df.columns = ['mcq','att']
            mcq_list = [x for x in df.mcq]
            att_list = [x for x in df.att]
            att_dict = dict(zip(mcq_list,att_list))
            expr = ','.join( [ f"mcq_attempts{x} = {att_dict[x]}"  for x in mcq_list] )
            updqry = "update userdata set " + expr + f" WHERE client_name = '{client_name}' AND courseid = '{course_id}' AND studentid={sid};"
            rds_update(updqry)

    syslog("reset mcq scores")
    updqry = "update userdata set " + ','.join([ "mcq_avg" + str(x) + " = 0"  for x in range( 1,max_mcq + 1 )]) + " where "
    updqry += condqry
    rds_update(updqry)

    userlist = list(set([x for x in mcq_df.student_id]))
    for sid in userlist:
        updqry = ""
        df1 = mcq_df[mcq_df.student_id==sid][['mcq','avgscore']].copy()
        mcq_list = list(set([x for x in df1.mcq]))
        for mcq in mcq_list:
            df2 = df1[ df1.mcq==mcq ]
            score = df2.avgscore.values[0]
            updqry += ",mcq_avg" + str(mcq) + " = " + str(score)
        updqry = "update userdata set " + updqry[1:] + " where studentid = " + str(sid) + " and " + condqry
        rds_update(updqry)
    syslog(f"completed on {course_id}")
    return

def edx_import(course_id,client_name,force_active=False):
    global max_iu
    [ pillar,course_code,module_code ] = course_header(client_name,course_id)
    if (pillar=="") or (course_code=="") or (module_code==""):
        syslog(f"The course setup for {course_id} is incomplete, edx_import stopped")
        return False
    cohort_id = getcohort(course_id)
    condqry = f"client_name = '{client_name}' and courseid = '{course_id}';"
    qry = f"select enabled from course_module where client_name = '{client_name}' and pillar = '{pillar}' and course_code = '{course_code}' and module_code = '{module_code}';"
    module_enabled = rds_param(qry)
    if module_enabled==0:
        syslog(f"this module {module_code} is not enabled, edx_import stopped")
        return False
    course_name = edx_coursename(course_id)
    if force_active:
        eoc = 0
    else:
        eoc = edx_endofcourse(client_name,course_id)
    update_playbooklist(course_id,client_name,course_name,eoc)
    if eoc==1:
        syslog(f"this course {course_id} already expired. No need to import again")
        return False

    cnt = rds_param("select count(*) as cnt from stages where client_name='{client_name}' and courseid='{course_id}';")
    cnt = int("0" + str(cnt))
    if cnt==0:
        qry = f"select * from stages_master where client_name = '{client_name}' and pillar = '{pillar}' and course_code = '{course_code}' and module_code = '{module_code}';"
        df = rds_df(qry)
        if df is not None:
            df.columns = get_columns("stages_master")
            query = "delete from stages where " + condqry
            rds_update(query)
            df['stagedate'] = ''
            df['courseid'] = course_id
            df1 = df[['client_name','courseid','id','stage','name','desc','days','f2f','mcq','flipclass','assignment','IU','stagedate']]
            copydbtbl(df1,"stages")

    df = edx_userdata(course_id)
    nrows = len(df)
    if nrows == 0:
        errlog("no user data from edx api found")
        return False
    query = "delete from userdata where " + condqry
    rds_update(query)
    query = f"SELECT `name` FROM stages WHERE client_name = '{client_name}' AND courseid='{course_id}' AND "
    query += "STR_TO_DATE(startdate,'%d/%m/%Y') <= CURDATE() ORDER BY id DESC LIMIT 1;"
    stage = rds_param(query)
    if stage=="":
        stage = "SOC Days"
    df['client_name'] = client_name
    df['amt'] = 0
    df['grade'] = 0
    df['stage'] = stage
    df['f2f'] = 0
    df['deferred'] = 0
    df['risk_level'] = 0
    df['mcq_zero'] = '[]'
    df['mcq_failed'] = '[]'
    df['as_zero'] = '[]'
    df['as_failed'] = '[]'
    df.rename(columns={'course_id':'courseid','student_id':'studentid'} ,inplace=True)
    df1 = df[['client_name','courseid','studentid','username','amt','grade','stage','f2f','deferred','risk_level','mcq_zero','mcq_failed','as_zero','as_failed']]
    copydbtbl(df1,"userdata")

    syslog("get # of mcq and assignment tests")
    mlist = []
    rng_iu = range( 1,max_iu + 1 )
    list1 = [ "mcq_avg"+ str(x) + " = 0" for x in rng_iu]
    list2 = [ "mcq_attempts"+ str(x) + " = 0" for x in rng_iu]
    list3 = [ "as_avg"+ str(x) + " = 0" for x in rng_iu]
    list4 = [ "as_attempts"+ str(x) + " = 0" for x in rng_iu]
    mlist += list1 + list2 + list3 + list4
    query = "update userdata set " + ','.join(mlist) + " where " + condqry
    try:
        rds_update(query)
    except:
        errlog("unable to initialized table userdata")
        return False

    update_assignment(course_id,client_name)
    update_mcq(course_id,client_name)
    update_schedule(course_id,client_name)
    return True

def count_avg_cols(client_name,course_id,maxmcqtest,colname):
    condqry = " from userdata where courseid = '" + course_id + "' and client_name = '" + client_name + "';"
    query =  'select ' + ','.join([ colname + str(x) for x in range( 1,maxmcqtest + 1 )]) + condqry
    df = rds_df(query)
    if df is None:
        return []
    rowcnt = len(df)
    if rowcnt==0:
        return []
    query =  'select ' + ','.join([ 'sum(' + colname + str(x) + ')/' + str(rowcnt) + ' as x' + str(x) for x in range( 1,maxmcqtest + 1 )])
    query +=  condqry
    df = rds_df(query)
    return [df[x][0] for x in list(df) ]

def gen_mcqas_df(colsum,client_name,course_id):
    if sum(colsum)==0:
        return None
    query = "select client_name, courseid as course_id, studentid, grade, 0 as mcq_sumscore, 0 as as_sumscore"
    xopts = ["mcq","as"]
    cols = ['client_name','course_id','studentid','grade','mcq_sumscore','as_sumscore']
    for xvar in xopts:
        n = xopts.index(xvar)
        rsum = colsum[n]
        if rsum==0:
            query += ",0 as " + xvar + "_maxattempts"
        else:
            xavg = xvar + "_avg"
            xatt = xvar + "_attempts"
            cols += [ xavg + str(x) for x in range( 1,rsum + 1 ) ]
            cols += [ xatt + str(x) for x in range( 1,rsum + 1 ) ]
            updqry1 = ''.join([ ', ' + xavg + str(x) for x in range( 1,rsum + 1 )])
            updqry2 = ''.join([ ', ' + xatt + str(x) for x in range( 1,rsum + 1 )])
            updqry3 = ', '.join([ xatt + str(x) for x in range( 1,rsum + 1 )])
            query += updqry1 + updqry2
    query += " from userdata where client_name = '" + client_name + "' and courseid = '" + course_id + "';"
    df = rds_df(query)
    if df is None:
        return
    df.columns = cols
    for xvar in xopts:
        df[xvar + "_maxattempts"] = 0
        df_expr = ', '.join([ "x." + xatt + str(x) for x in range( 1,rsum + 1 ) ])
        df_expr = "max(" + df_expr + ")"
        df[xvar + "_maxattempts"] = df.apply(lambda x : eval(df_expr),axis=1)
    for xvar in xopts:
        try:
            n = xopts.index(xvar)
            rsum = colsum[n]
            if rsum==0:
                df[xvar + '_cnt'] = 0
                df[xvar + "_avgattempts"] = 0
                df[xvar + '_avgscore'] = 0
            else:
                xavg = xvar + "_avg"
                xatt = xvar + "_attempts"
                cols1 = [ xavg + str(x) for x in range( 1,rsum + 1 )]
                df[ xvar + "_sumscore" ] = df[cols1].sum(axis=1)
                cols = [ xatt + str(x) for x in range( 1,rsum + 1 )]
                df[ xvar + "_sumattempts" ] = df[cols].sum(axis=1)
                for n in range(1,rsum + 1):
                    attflag = xatt + str(n)
                df[attflag] = df[attflag].apply(lambda x: 1 if x>0 else 0)
                df[xvar + '_cnt'] = df.apply(lambda x: sum([ 1 if x[c]>0 else 0 for c in cols1 ]),axis=1)
                df[xvar + "_avgattempts"] = df.apply(lambda x: int(x[xvar + "_sumattempts"]*100/ x[xvar + "_cnt"])/100 if x[xvar + "_cnt"] > 0 else 0,axis=1)
                df[xvar + '_avgscore'] = df.apply(lambda x: int(x[xvar + "_sumscore"]*100/ x[xvar + "_cnt"])/100 if x[xvar + "_cnt"] > 0 else 0,axis=1)
        except:
            errlog(f"error updating {xvar}")

    try:
        df1 = df[ (df.mcq_cnt + df.as_cnt) > 0 ]
        df1 = df1[["client_name","course_id","studentid","grade","mcq_avgscore","mcq_avgattempts", \
                  "mcq_maxattempts","mcq_cnt","as_avgscore","as_avgattempts","as_maxattempts","as_cnt"]]
        return df1
    except:
        return None

def update_mcqas_info(client_name,course_id):
    global max_iu
    query = f"delete from mcqas_info where client_name = '{client_name}' and course_id = '{course_id}';"
    rds_update(query)

    colsum=[]
    for xvar in ["mcq_avg","as_avg"]:
        avgsum_list = count_avg_cols(client_name,course_id,max_iu,xvar)
        if len(avgsum_list)>0:
            rr = [ 1 if r>0 else 0 for r in avgsum_list ]
            rsum = sum(rr)
            colsum.append(rsum)

    if len(colsum) > 0:
        df = gen_mcqas_df(colsum,client_name,course_id)
        if df is None:
            syslog(f"there is no data for {course_id}")
            return
        copydbtbl(df,"mcqas_info")
    syslog(f"completed on {course_id}")
    return

def generate_mcq_as(client_name):
    df = rds_df("select distinct course_id from playbooks where client_name = '" + client_name + "';")
    df.columns = ['course_id']
    course_list = [x for x in df.course_id]
    syslog(f"Generating data based on playbooks")
    for course_id in course_list:
        try:
            update_mcqas_info(client_name,course_id)
        except:
            pass
    syslog(f"completed on {client_name}")
    return

def update_userdf(userdf,client_name):
    query = f"select distinct studentid from user_master where client_name = '{client_name}' order by studentid;"
    df = rds_df(query)
    if df is None:
        sid_list = []
    else:
        df.columns = ['studentid']
        sid_list = [ x for x in df.studentid]
    df = userdf
    df.rename(columns={'course_id':'courseid','student_id':'studentid'},inplace=True)
    df['client_name'] = client_name
    df['courseid'] = ''
    df['chat_id'] = 0
    df['binded'] = 0
    df['usertype'] = 1
    df['existed'] = df.apply(lambda x: 1 if x['studentid'] in sid_list else 0,axis=1)
    df = df[df.existed == 0]
    df = df[['client_name','studentid','username','email','usertype','binded','chat_id','courseid']]
    if len(df)==0:
        return df
    email_filter = rds_param(f"SELECT `value` from params WHERE  `key` = 'email_filter' and client_name = '{client_name}';")
    efilter = email_filter.split(',')
    df['usertype'] = df.apply(lambda x: 11 if x['email'].lower().split('@')[1] in efilter else 1,axis=1)
    return df

def get_calendar_json(api_url):
    data = {}
    syslog(f"calling api {api_url} via requests.post")
    response = requests.get(api_url)
    syslog(f"completed with status code {response.status_code}")
    if response.status_code==200:
        result = response.content.decode('utf-8')
        if len(result)>0:
            data = json.loads(result)
    return data

def get_stage_list(data,module_node):
    if data == {}:
        return []
    resp = data['events']
    if len(resp)==0:
        return []
    stage_list = []
    soc_date = resp[0]['startDate'][:10]
    socd = dmy_str(soc_date)
    socdt = ymd_str(soc_date)
    s_dict = {}
    s_list = []
    stg_id = 0
    for x in resp:
        if 'summary' not in list(x):
            syslog("summary element missing")
            continue
        if 'description' not in list(x):
            syslog("description element missing")
            continue
        x_summary = x['summary']
        x_desc = x['description']
        if x_summary is None:
            x_list=[]
            cohort = ""
            stage_name = ""
            stage_desc = ""
        else:
            x_list = [x.strip() for x in x_summary.split(':')]
            n=len(x_list)
            cohort = x_list[0] if n>0 else ""
            stage_name = x_list[1] if n>2 else ""
            stage_desc = x_list[2] if n>2 else stage_name
            stage_desc = stage_desc.replace('?','').split('(')[0]
        cohort = cohort.strip()
        iu_list = "0"
        stdate = ymd_str(x['startDate'][:10])
        sdate = dmy_str(x['startDate'][:10])
        edate = dmy_str(x['endDate'][:10])
        stage_name = stage_code(stage_name)
        if (stage_name != "") and (cohort==module_node) :
            stg_id += 1
            stg_key = str(stdate) + "_" + str(stg_id)
            stage_info = [stage_name,cohort,stage_desc,sdate,edate,iu_list]
            s_dict[stg_key] = stage_info
            s_list.append(stg_key)
    if s_list==[]:
        syslog("stages info s_list empty")
        return []
    s_list.sort()
    s_list = sorted(s_list)
    n = 0
    cohort_dict = {}
    index_dict = {}
    cohort_list = []
    for x in [ s_dict[x][1] for x in s_list ]:
        if x not in cohort_list:
            cohort_list.append(x)
    for x in cohort_list:
        cohort_dict[x] = {}
        index_dict[x] = []
    for x in s_list:
        y = s_dict[x][1]
        w = x.split('_')[0]
        cohort_dict[y][x]=s_dict[x]
        index_dict[y].append(x)
    for x in cohort_list:
        z = cohort_dict[x]
        w = [ vv[0] for (kk,vv) in z.items() ]
        dt0 = min([ kk for (kk,vv) in z.items() ])
        startdt = cohort_dict[x][dt0][3]
        dt0 = str2date(startdt) - datetime.timedelta(days=1)
        if 'SOC' not in w:
            dt0a = ymd_str(str(dt0))
            dt0b = dmy_str(str(dt0))
            cohort_dict[x][dt0a] = ["SOC","SOC Days","Start of course",dt0b,startdt,"0"]
            index_dict[x].append(dt0a)
        if 'EOC' not in w:
            dt1 = max([ kk for (kk,vv) in z.items() ])
            enddt = cohort_dict[x][dt1][3]
            dt1 = str2date(enddt) + datetime.timedelta(days=1)
            dt1a = ymd_str(str(dt1))
            dt1b = dmy_str(str(dt1))
            cohort_dict[x][dt1a] = ["EOC",x,"End of course",enddt,dt1b,"0"]
            index_dict[x].append(dt1a)
        s_list = sorted(index_dict[x])
        index_dict[x] = s_list
        stg = "SOC"
        sorted_stage_list = []
        for y in list(index_dict[x]):
            if cohort_dict[x][y][0] == "SOC":
                cohort_dict[x][y][1] = "SOC Days"
            else:
                if stg == cohort_dict[x][y][0]:
                    cohort_dict[x][y][0] += "2"
                    cohort_dict[x][y][2] += " 2"
                stg = stg + ' - ' + cohort_dict[x][y][0]
                cohort_dict[x][y][1] = stg
            stg = cohort_dict[x][y][0]
            cohort_dict[x][y].append((str2date(cohort_dict[x][y][3]) - dt0).days )
            sorted_stage_list.append(cohort_dict[x][y])
        cohort_dict[x] = sorted_stage_list
    return cohort_dict

def stage_code(txt):
    tt = [t for t in txt]
    n=len(txt)
    wt = ''.join([ ' ' if tt[i-1].isnumeric() and not tt[i].isnumeric() else tt[i] for i in range(1,n)])
    wlist = ['learning','assignment','flipped','implementation','mentoring','assessment','orientation']
    vlist = ['EL','A','FC','PI','PM','SA','SOC']
    ulist = [ w for w in wt.split(' ') if w!='' and w.isnumeric() ]
    uu = ulist[0] if len(ulist)>0 else ''
    tt = txt.lower()
    result = [w for w in wlist if w in tt]
    stg = ""
    if result == []:
        result = [ x for x in vlist if txt.startswith(x) ]
        if result == []:
            return ""
        stg = txt
    else:
        n = wlist.index(result[0])
        stg = vlist[n] + uu
    return stg

def get_google_calendar(course_id,client_name):
    qry = f"select cohort_id from playbooks where client_name = '{client_name}' and course_id = '{course_id}';"
    [ pillar,course_code,module_code ] = course_header(client_name,course_id)
    cohort_id = getcohort(course_id)
    api_url = f"https://realtime.sambaash.com/v1/calendar/fetch?cohortId={course_code}%20:%20{cohort_id}"
    data  = get_calendar_json(api_url)
    if data == {}:
        syslog("there is no data from google calendar")
        return []
    sorted_stage_list = get_stage_list(data,cohort_id)
    stage_list = []
    for cohort in list(sorted_stage_list) :
        stage_list = sorted_stage_list[cohort]
    return stage_list

def update_stage_table(stage_list,course_id,client_name,cohort_id,pillar,course_code,module_code):
    qry = " client_name = '_c_' and courseid = '_x_';"
    qry = qry.replace('_c_',client_name)
    qry = qry.replace('_x_',course_id)
    mcq_dict = {}
    ast_dict = {}
    stg_list = []
    fc_dict = {}
    iu_dict = {}
    f2f_dict = {}

    query = "select * from stages where " + qry
    df = rds_df(query)
    if (df is None) or (len(stage_list)==0):
        query = f"select * from stages_master where client_name = '{client_name}' and module_code = '{module_code}'"
        if (pillar != "") and (course_code != ""):
            query += f" and course_code='{course_code}' and pillar='{pillar}';"
        df = rds_df(query)
        if df is None:
            return []
        df.columns = get_columns("stages_master")
        stg_list = [x for x in df.stage]
        stgname_list = [x for x in df.name]
        stgdesc_list = [x for x in df.desc]
        stgdays_list = [x for x in df.days]
        mcq_list = [x for x in df.mcq]
        ast_list = [x for x in df.assignment]
        fc_list = [x for x in df.flipclass]
        iu_list = [x for x in df.IU]
        f2f_list = [x for x in df.f2f]
        mcq_dict = dict(zip(stg_list,mcq_list))
        ast_dict = dict(zip(stg_list,ast_list))
        fc_dict = dict(zip(stg_list,fc_list))
        iu_dict = dict(zip(stg_list,iu_list))
        f2f_dict = dict(zip(stg_list,f2f_list))
        dt0 = edx_day0(course_id)
        stage_list = []
        id = 0
        cnt = len(stg_list)
        for n in range(cnt):
            fld1 = stg_list[n]
            fld2 = stgname_list[n]
            fld3 = stgdesc_list[n]
            m = n if (cnt == n + 1) else n+1
            fld7 = stgdays_list[n]
            dd = stgdays_list[m]
            dt1 = dt0 + datetime.timedelta( days = (int(fld7)-1) )
            dt2 = dt0 + datetime.timedelta( days = (int(dd)-1) )
            fld4 = dt1.strftime('%d/%m/%Y')
            fld5 = dt2.strftime('%d/%m/%Y')
            fld6 = iu_list[n]
            stage_list.append([fld1,fld2,fld3,fld4,fld5,fld6,fld7])
    else:
        df.columns = get_columns("stages")
        stg_list = [x for x in df.stage]
        mcq_list = [x for x in df.mcq]
        ast_list = [x for x in df.assignment]
        fc_list = [x for x in df.flipclass]
        iu_list = [x for x in df.IU]
        f2f_list = [x for x in df.f2f]
        mcq_dict = dict(zip(stg_list,mcq_list))
        ast_dict = dict(zip(stg_list,ast_list))
        fc_dict = dict(zip(stg_list,fc_list))
        iu_dict = dict(zip(stg_list,iu_list))
        f2f_dict = dict(zip(stg_list,f2f_list))

    m = len(stage_list)
    if m==0:
        return []

    for x in stage_list:
        if x[0] not in stg_list:
            mcq_dict[ x[0] ] = "0"
            ast_dict[ x[0] ] = "0"
            fc_dict[ x[0] ] = "0"
            iu_dict[ x[0] ] = "0"
            f2f_dict[ x[0] ] = 0
    for x in stage_list:
        x.append( mcq_dict[x[0]] )
        x.append( ast_dict[x[0]] )
        x.append( fc_dict[x[0]] )
        x.append( iu_dict[x[0]] )
        x.append( f2f_dict[x[0]] )

    query = "delete from stages where " + qry
    rds_update(query)

    n = 0
    for n in range(len(stage_list)):
        stg = stage_list[n][0]
        query = "insert into stages(client_name,courseid,id,stage,`name`,`desc`,days,f2f,mcq,assignment,IU,flipclass) values("
        query += "'" + client_name + "','" + course_id + "'," + str(n+1) # client_name,courseid,id
        query += ",'" + str(stage_list[n][0]) + "','"  # stage
        query += str(stage_list[n][1]) + "','"         # stage_name
        query += str(stage_list[n][2]) + "',"          # stage_desc
        query += str(stage_list[n][6]) + ","           # days
        query += str(stage_list[n][11]) + ",'"         # f2f
        query += str(stage_list[n][7]) + "','"         # mcq
        query += str(stage_list[n][8]) + "','"         # assignment
        query += str(stage_list[n][10])  + "','"       # IU
        query += str(stage_list[n][9]) + "')"          # flipclass
        rds_update(query)

    arr_stagedate = [ x[3] for x in stage_list]
    m = len(arr_stagedate)
    if m==0:
        syslog("There is not enough information to build the schedule")
        return []
    for n in range(m - 1):
        id = n + 1
        start_date = arr_stagedate[n]
        stg_date = arr_stagedate[id]
        query = f"update stages set startdate = '{start_date}', stagedate = '{stg_date}' where id = {id} and " + qry
        try:
            rds_update(query)
        except:
            syslog("error in sql update\n{query}")
            pass

    start_date = arr_stagedate[-1]
    eoc_date = string2date(stage_list[-1][4],"%d/%m/%Y")
    stg_date = eoc_date.strftime('%d/%m/%Y')
    query = f"update stages set startdate = '{start_date}',stagedate = '{stg_date}' where id = {m} and " + qry
    rds_update(query)
    query = f"SELECT `name` FROM stages WHERE client_name = '{client_name}' AND courseid='{course_id}' AND "
    query += "STR_TO_DATE(stagedate,'%d/%m/%Y') <= CURDATE() ORDER BY id DESC LIMIT 1;"
    stage = rds_param(query)
    if stage != "":
        query=f"update userdata set stage='{stage}' WHERE client_name = '{client_name}' AND courseid='{course_id}' ;"
        try:
            rds_update(query)
        except:
            syslog("error in sql update\n{query}")
    return stage_list

def update_playbooklist(course_id,client_name,course_name,eoc):
    query = f"select count(*) as cnt from playbooks where client_name = '{client_name}' and course_id = '{course_id}';"
    cnt = rds_param(query)
    if cnt==1:
        return
    [ pillar,course_code,module_code ] = course_header(client_name,course_id)
    cohort_id = getcohort(course_id)
    query = "insert into playbooks(client_name,pillar,course_code,module_code,cohort_id,course_id,course_name,mentor,eoc) \
             values('_c_','_p_','_w_','_m_','_x_','_y_','_z_','',_e_);"
    query = query.replace("_c_",client_name)
    query = query.replace("_p_",pillar)
    query = query.replace("_w_",course_code)
    query = query.replace("_m_",module_code)
    query = query.replace("_x_",cohort_id)
    query = query.replace("_y_",course_id)
    query = query.replace("_z_",course_name)
    query = query.replace("_e_",str(eoc))
    rds_update(query)
    return

def edx_mass_update(func,clt):
    global edx_api_header,edx_api_url
    module_code = lambda x : piece(piece(piece(x,':',1),'+',1),'-',0)
    date_today = datetime.datetime.now().date()
    keyword = date_today.strftime('%Y')[-2:]
    if clt=="":
        with open("vmbot.json") as json_file:
                bot_info = json.load(json_file)
        client_name = bot_info['client_name']
    else:
        client_name = clt

    qry = f"delete from userdata where client_name='{client_name}' and courseid = '';"
    rds_update(qry)
    course_list = search_course_list(keyword)
    if course_list==[]:
        syslog("there is nothing found, check the rest api.")
        return
    qry = f"SELECT DISTINCT module_code FROM course_module WHERE enabled=1 and client_name='{client_name}';"
    df = rds_df(qry)
    if df is None:
        mc_list = []
    else:
        df.columns = ['module_code']
        mc_list = [x for x in df.module_code]
    qry = f"select course_id, module_code from playbooks where client_name = '{client_name}'"
    df = rds_df(qry)
    if df is None:
        c_list = []
        m_list = []
        mc_dict = {}
    else:
        df.columns = ['course_list','module_code']
        c_list = [x for x in df.course_list]
        m_list = [x for x in df.module_code]
        mc_dict = dict(zip(c_list,m_list))
    for course_id in course_list:
        if course_id in c_list:
            mcode = mc_dict[course_id]
        else:
            mcode = module_code(course_id)
        if mcode in mc_list:
            eoc = edx_endofcourse(client_name,course_id)
            query = f"update playbooks set eoc = {eoc} where client_name = '{client_name}' and course_id = '{course_id}';"
            rds_update(query)
            if eoc == 0:
                func(course_id,client_name)
    return

def edx_mass_import(client_name):
    edx_mass_update(edx_import,client_name)
    return

def mass_update_assignment(client_name):
    edx_mass_update(update_assignment,client_name)
    return

def mass_update_mcq(client_name):
    edx_mass_update(update_mcq,client_name)
    return

def mass_update_schedule(client_name):
    edx_mass_update(update_schedule,client_name)
    return

def mass_update_usermaster(client_name):
    userdf = edx_alluserdata()
    if len(userdf)==0:
        return
    df = update_userdf(userdf,client_name)
    if df is None:
        syslog("Failed to perform user master update")
        return
    if len(df)==0:
        syslog("nothing to update")
        return
    copydbtbl(df,"user_master")
    syslog(f"completed on {client_name}")
    return

def edx_alluserdata():
    global edx_api_header,edx_api_url
    df = pd.DataFrame.from_dict({'student_id':[],'course_id':[],'username':[],'email':[]})
    url = f"{edx_api_url}/user/fetch/all"
    syslog(f"calling api {url} via requests.post")
    response = requests.get(url,headers=edx_api_header)
    syslog(f"completed with status code {response.status_code}")
    if response.status_code==200:
        data = response.content.decode('utf-8')
        if str(data) == "":
            return None
        userinfo = json.loads(data)
        rows_cnt = len(list(userinfo))
        if rows_cnt == 0:
            return df
        user_list = []
        course_list = []
        username_list = []
        email_list = []
        for usr in userinfo:
            uid = usr['user_id']
            course_id = usr['enrolments'][0]['course_id']
            user_list.append(uid)
            course_list.append(course_id)
            username_list.append(usr['username'])
            email_list.append(usr['email'])
        data = {'student_id':user_list,'course_id':course_list,'username':username_list,'email':email_list}
        df = pd.DataFrame.from_dict(data)
    return df

def load_edx(client_name):
    query = f"select `value` from params WHERE client_name='{client_name}' AND `key`='email_filter';"
    email_filter = rds_param(query)
    efilter = email_filter.split(',')
    date_today = datetime.datetime.now().date()
    yr_str = str(date_today.strftime('%Y'))
    yrnow = yr_str[-2:]
    query = f"SELECT DISTINCT u.courseid FROM userdata u INNER JOIN playbooks p "
    query += f" ON u.client_name=p.client_name AND u.courseid=p.course_id "
    query += f"WHERE u.client_name = '{client_name}' AND p.eoc=0;"
    df = rds_df(query)
    if df is None:
        course_list = []
    else:
        df.columns = ['courseid']
        course_list = [x for x in df.courseid]

    updated_courses = []
    vars = dict()
    txt =  ""
    for course_id in course_list:
        eoc = edx_endofcourse(client_name,course_id)
        eoc_gap = edx_eocgap(client_name,course_id,7)
        if (eoc == 0) or (eoc_gap==1):
            update_mcq(course_id,client_name)
            update_assignment(course_id,client_name)
            update_schedule(course_id,client_name)
            updated_courses.append(course_id)
        if eoc == 1:
            query = f"update playbooks set eoc=1 where client_name='{client_name}' AND course_id='{course_id}';"
            rds_update(query)
    query = f"SELECT DISTINCT u.courseid FROM userdata u INNER JOIN playbooks p "
    query += f" ON u.client_name=p.client_name AND u.courseid=p.course_id "
    query += f"WHERE u.client_name = '{client_name}' AND p.eoc=1 AND SUBSTRING(u.courseid,-2)='{yrnow}';"
    df = rds_df(query)
    if df is None:
        eoc_list = []
    else:
        df.columns = ['courseid']
        eoc_list = [x for x in df.courseid]
    course_list = search_course_list(yrnow)
    course_list = [x for x in course_list if (x[-3].isnumeric()==False and x[-2:]==yrnow) or x[-4:]==yr_str ]
    for course_id in [ x for x in course_list if x not in updated_courses and x not in eoc_list]:
        edx_import(course_id,client_name)
        updated_courses.append(course_id)
    mass_update_usermaster(client_name)
    query = f"UPDATE user_master SET usertype=0 WHERE client_name = '{client_name}' AND studentid IN "
    query += " ( SELECT studentid FROM ( SELECT b.studentid, SUM( IFNULL((SELECT 1 FROM playbooks WHERE client_name=b.client_name AND "
    query += " course_id=b.courseid AND eoc=0),0)) AS cnt FROM userdata b INNER JOIN user_master c ON b.client_name=c.client_name AND "
    query += f" b.studentid=c.studentid WHERE b.client_name = '{client_name}' AND c.usertype=1 group by b.studentid ) zz WHERE cnt = 0);"
    rds_update(query)
    for tbl in ['stages_master','stages']:
        for fld in ['mcq','assignment','IU']:
            query = f"UPDATE {tbl} SET {fld}='0' WHERE client_name = '{client_name}' and {fld} = '';"
            rds_update(query)
    return

def list_tables(module_code,details):
    numstr = lambda x : int(x) if str(x).isnumeric else 0
    cnt = len(details[0])
    iu_cnt = [ n for n in range(cnt) if details[0][n]=='' ][0] - 1
    list_iunum = [ numstr(details[0][n]) for n in range(1,iu_cnt+1)]
    list_lm = [ details[1][n] for n in range(1,iu_cnt+1)]
    list_mcqnum = [ numstr(details[2][n]) for n in range(1,iu_cnt+1)]
    list_numqns = [ numstr(details[3][n]) for n in range(1,iu_cnt+1)]
    list_asnum = [ numstr(details[4][n]) for n in range(1,iu_cnt+1)]
    as_cnt = sum(list_asnum)
    list_pm = [ numstr(details[5][n]) for n in range(1,iu_cnt+1)]
    list_pi = [ numstr(details[6][n]) for n in range(1,iu_cnt+1)]
    module_iu = {'iu': list_iunum,'desc': list_lm,'mcq': list_mcqnum,'qn': list_numqns,'as': list_asnum,'pm': list_pm,'pi': list_pi}
    list1 = [ n for n in range(cnt) if n > iu_cnt if details[0][n] != '']
    m = iu_cnt if list1==[] else list1[0]
    list2 = [ n for n in range(cnt) if (n > m) and (details[0][n]=='')]
    p = cnt+1 if list2==[] else list2[0]
    rownum_list = [ n for n in range(cnt) if (n > m) and (n<p) ]
    list1 = [ d[m] for d in details if d[m] != '' ]
    cnt = len(list1)
    col_dict=dict()
    for n in range(cnt):
        col_dict[ list1[n] ] = n
    list_stage_id = [ n+1 for n in range(len(rownum_list)) ]
    list_stage = [ details[col_dict['stage']][n] for n in rownum_list ]
    list_name = [ 'SOC Days' if n==0 else list_stage[n-1] + ' - ' + list_stage[n] for n in range(len(rownum_list)) ]
    list_desc = [ details[col_dict['desc']][n] for n in rownum_list ]
    list_days = [ details[col_dict['days']][n] for n in rownum_list ]
    list_mcq = [ details[col_dict['mcq']][n] for n in rownum_list ]
    list_mcq = [ '0' if str(x)=='' else str(x) for x in list_mcq]
    list_mcq = [ ','.join([x for x in y.split(',') if int(x) <= iu_cnt]) for y in list_mcq ]
    list_assignment  = [ details[col_dict['assignment']][n] for n in rownum_list ]
    list_assignment = [ '0' if str(x)=='' else str(x) for x in list_assignment]
    list_assignment = [ ','.join([x for x in y.split(',') if int(x) <= iu_cnt]) for y in list_assignment ]
    list_iu = [ details[col_dict['IU']][n] for n in rownum_list ]
    list_iu = [ '0' if str(x)=='' else str(x) for x in list_iu]
    stages_master = {'id': list_stage_id,'stage': list_stage,'name': list_name,'desc': list_desc,'days': list_days,'mcq': list_mcq,'assignment': list_assignment,'iu': list_iu}
    return ( module_iu,stages_master )

def valid_format(original_fn,fn):
    try:
        wb = load_workbook(fn)
    except:
        errmsg = f"valid_format : Unable to load the file {fn}"
        return (False,errmsg,None)
    list_sheets = wb.sheetnames
    if len(list_sheets) <= 1:
        errmsg = "This workbook is should have at least 2 worksheets inside !"
        return (False,errmsg,None)
    if 'Master Course Details' not in list_sheets:
        errmsg = "Missing 'Master Course Details' sheet in this file."
        return (False,errmsg,None)
    exclusion_list = ['Intro Sheet','Master Course Details','Cohorts','course_id']
    active_modules = [ x for x in list_sheets if x not in exclusion_list]
    ws = wb['Master Course Details']
    master = [ ['' if ws.cell(m+1,n+1).value is None else ws.cell(m+1,n+1).value for m in range(ws.max_row)] for n in range(ws.max_column)]
    m=min([len(x) for x in master])
    if (len(master) < 12) or (m <= 1):
        errmsg = "The worksheet 'Master Course Details' not in correct format"
        return (False,errmsg,None)
    cnt = len(master[5])
    if cnt <= 1:
        errmsg = "The module code list in the worksheet 'Master Course Details' is empty."
        return (False,errmsg,None)
    pillar = master[0][1]
    course_code = master[3][1]
    n = len(pillar)
    pillar_code = original_fn[:n].upper()
    if (pillar_code != pillar) or (n==0):
        errmsg = f"Unable to import from this file {original_fn} due to pillar code {pillar} mismatch with the filename prefix"
        return (False,errmsg,None)
    return (True,wb,master)

def course_import(client_name,fn,wb,master):
    def is_valid_courseid(course_id, mcode_list):
        valid_courseid = False
        for module_code in mcode_list:
            mcode = "+" + module_code
            if mcode in course_id:
                valid_courseid = True
                break
        return valid_courseid
    list_sheets = wb.sheetnames
    pillar = master[0][1]
    course_code = master[3][1]
    course_provider_partner = master[1][1]
    course_name = master[4][1]
    blended_learning_e = master[7][1]
    calendar = master[9][1]
    title = f"Pillar = {pillar} Course code = {course_code}"
    course_dict = dict()
    iucnt_dict = dict()
    mcq_cnt_dict = dict()
    as_cnt_dict = dict()
    stages_cnt_dict = dict()
    module_code_list = [x for x in master[5] if x != '']
    cnt = len(module_code_list)
    df = None
    for n in range(1,cnt):
        module_code = master[5][n]
        module_name = master[6][n]
        course_dict[module_code] = module_name

    active_modules = list(course_dict)
    import_list = [ x for x in list(list_sheets ) if x in active_modules ]
    where_cond = f"where client_name = '{client_name}' and pillar = '{pillar}' and course_code='{course_code}' and module_code in ('"
    where_cond += "','".join(import_list) + "');"
    tbl_list = ['module_iu','stages_master','course_module']
    for tbl in tbl_list:
        sql_code = f"delete from {tbl} " + where_cond
        vmsvclib.rds_update(sql_code)
    update_sql = "INSERT INTO course_module (`client_name`, `pillar`, `course_code`, `module_code`, `module_name`, `mcq_cnt`, `as_cnt`) VALUES "
    for module_code in import_list:
        module_name = course_dict[module_code]
        ws = wb[module_code]
        details = [ ['' if ws.cell(m+1,n+1).value is None else ws.cell(m+1,n+1).value for m in range(ws.max_row)] for n in range(ws.max_column)]
        (module_iu,stages_master) = list_tables(module_code,details)
        iu_cnt = len(module_iu['iu'])
        mcq_cnt = sum(module_iu['mcq'])
        as_cnt = sum(module_iu['as'])
        sql_hdr = f"('{client_name}','{pillar}','{course_code}','{module_code}'"
        sql_code = "INSERT INTO `module_iu` (`client_name`, `pillar`, `course_code`, `module_code`, `iu`, `learning_materials`, `mcq`, `number_of_qns`, `assignment`, `flipped_class`, `assignment_support`, `project_mentoring`, `project_implementation`) VALUES "
        for n in range(iu_cnt):
            sep = ";" if (iu_cnt - n) == 1 else ","
            sql_code += f"{sql_hdr},{module_iu['iu'][n]},'{module_iu['desc'][n]}',{module_iu['mcq'][n]},{module_iu['qn'][n]},"
            sql_code += f"{module_iu['as'][n]},'0','0',{module_iu['pm'][n]},{module_iu['pi'][n]}){sep}"
        vmsvclib.rds_update(sql_code)
        stage_cnt = len(stages_master['id'])
        iucnt_dict[module_code] = iu_cnt
        mcq_cnt_dict[module_code] = mcq_cnt
        as_cnt_dict[module_code] = as_cnt
        stages_cnt_dict[module_code] = stage_cnt
        sql_code = "INSERT INTO `stages_master` (`client_name`, `pillar` , `course_code` , `module_code`, `id`, `stage`, `name`, `desc`, `days`, `f2f`, `mcq`, `flipclass`, `assignment`, `IU`) VALUES "
        for n in range(stage_cnt):
            sep = ";" if (stage_cnt - n) == 1 else ","
            sql_code += f"{sql_hdr},{stages_master['id'][n]},'{stages_master['stage'][n]}',"
            sql_code += f"'{stages_master['name'][n]}','{stages_master['desc'][n]}','{stages_master['days'][n]}',0,'{stages_master['mcq'][n]}',"
            sql_code += f"'0','{stages_master['assignment'][n]}','{stages_master['iu'][n]}'){sep}"
        vmsvclib.rds_update(sql_code)
        sep = ';' if module_code==import_list[-1] else ','
        update_sql += f"{sql_hdr}, '{module_name}', {mcq_cnt}, {as_cnt}){sep}"
    vmsvclib.rds_update(update_sql)
    update_sql = f"update course_module set course_name='{course_name}',course_provider_partner='{course_provider_partner}',blended_learning_e='{blended_learning_e}', "
    update_sql += f"calendar='{calendar}',feedback_survey=0,mentor_email='',assistance_email='',enquiry_email='',enabled=1 "
    update_sql += where_cond
    vmsvclib.rds_update(update_sql)
    df = pd.DataFrame.from_dict({'module':list(import_list),\
            'IUs':[iucnt_dict[x] for x in import_list], \
            'MCQs':[mcq_cnt_dict[x] for x in import_list],\
            'Assignments':[as_cnt_dict[x] for x in import_list],\
            'Stages':[stages_cnt_dict[x] for x in import_list] })
    if 'course_id' in list_sheets:
        ws = wb['course_id']
        eoc = 0
        mentor = ''
        course_list = [ ['' if ws.cell(m+1,n+1).value is None else ws.cell(m+1,n+1).value for m in range(ws.max_row)] for n in range(ws.max_column)]
        if (len(course_list) >= 1) and (course_list[0][0]=="course_id"):
            for course_id in course_list[0][1:]:
                if is_valid_courseid(course_id, module_code_list[1:]):
                    cohort_id = getcohort(course_id)
                    module_code = getmodulecode(course_id)
                    course_name = course_dict[module_code] if module_code in list(course_dict) else ''
                    sql_code = f"DELETE FROM playbooks WHERE client_name='{client_name}' and course_id='{course_id}';"
                    vmsvclib.rds_update(sql_code)
                    sql_code = "INSERT INTO playbooks (`client_name`, `course_id`, `pillar`, `course_code`, `module_code`, `course_name`, `cohort_id`, `eoc`, `mentor`) VALUES "
                    sql_code += f"('{client_name}','{course_id}','{pillar}','{course_code}','{module_code}','{course_name}','{cohort_id}','{eoc}','{mentor}');"
                    vmsvclib.rds_update(sql_code)
    return (True,title,df)

def parse_commands(client_name,cmd,resp,arg):
    cmd_log = ""
    edxcmd_list = ['import','mcq','assignment','schedule','stages','calendar','info']
    if (cmd in edxcmd_list) and (resp != ""):
        sid = 0
        if (arg != "") and arg.isnumeric():
            sid = int(arg)
        if resp == '*':
            course_id = resp
            course_list = [resp]
        else:
            course_list = search_course_list("%" + resp + "%" )
        if cmd=="search":
            for course_id in course_list:
                cmd_log += str(course_id) + '\n'
        elif len(course_list)==1:
            course_id = course_list[0]
            cmd_log += f"{course_id}\t{client_name}\n"
            if cmd=="import":
                if course_id=='*':
                    edx_mass_import(client_name)
                    cmd_log += "Edx mass import completed\n"
                else:
                    if edx_import(course_id,client_name):
                        cmd_log += "Edx import completed\n"
                    else:
                        cmd_log += "Edx import not successful\n"
            elif cmd=="mcq":
                if course_id=='*':
                    mass_update_mcq(client_name)
                    cmd_log += "MCQ mass update completed\n"
                elif sid > 0 :
                    update_mcq(course_id,client_name,sid)
                    cmd_log += f"MCQ update for studentid {sid} completed\n"
                else:
                    update_mcq(course_id,client_name)
                    cmd_log += "MCQ update completed\n"
            elif cmd=="assignment":
                if course_id=='*':
                    mass_update_assignment(client_name)
                    cmd_log += "Assignment mass update completed\n"
                elif sid > 0 :
                    update_assignment(course_id,client_name,sid)
                    cmd_log += f"Assignment update for studentid {sid} completed\n"
                else:
                    update_assignment(course_id,client_name)
                    cmd_log += "Assignment update completed\n"
            elif cmd=="schedule":
                if course_id=='*':
                    mass_update_schedule(client_name)
                    cmd_log += "Schedule mass update completed\n"
                else:
                    if update_schedule(course_id,client_name):
                        cmd_log += "Schedule update completed\n"
                    else:
                        cmd_log += "Schedule update not successful\n"
            elif cmd=="stages":
                stage_list = get_google_calendar(course_id,client_name)
                for x in stage_list:
                    cmd_log += str(x) + '\n'
                cmd_log += "google calendar test completed\n"
            elif cmd=="calendar":
                df = rds_df(f"select * from stages where client_name='{client_name}' and courseid='{course_id}';")
                if df is not None:
                    df.columns = get_columns("stages")
                    cmd_log += vmbotlib.stage_calendar(df) + "\n"
            elif cmd=="info":
                dt = edx_day0(course_id)
                cmd_log += f"start date = {dt}\n"
                dt = eoc_date(client_name,course_id)
                cmd_log += f"EOC date = {dt}\n"
                eoc = edx_endofcourse(client_name,course_id)
                results = f"End of course (1:Yes,0:No) = {eoc}"
                [ pillar,course_code,module_code ] = course_header(client_name,course_id)
                cohort_id = getcohort(course_id)
                cmd_log += str( [ pillar,course_code,module_code,cohort_id ] ) + '\n'
                cmd_log += str(results) + '\n'
                df = edx_userdata(course_id)
                if df is not None:
                    sid_list = [x for x in df.student_id]
                    uname_list = [x for x in df.username]
                    udict = dict(zip(sid_list,uname_list))
                    cmd_log += ''.join([ str(k) + '\t' + v + '\n' for k,v in udict.items()])
                else:
                    cmd_log += "No students information found\n"
            elif cmd=="attendance":
                df = sms_df(course_id,sid)
                if df is not None and len(df)>0:
                    lecturer = df.account_name.values[0]
                    class_code = df.class_code.values[0]
                    cohort_code = df.class_cohort_code.values[0]
                    remarks = df.class_remarks.values[0]
                    txt = ""
                    txt += f"Lecturer = {lecturer}\n"
                    txt += f"class_code = {class_code}\n"
                    txt += f"class_cohort_code = {cohort_code}\n"
                    txt += f"class_remarks = {remarks}\n"
                    date_list = sms_datelist(df)
                    dlist = [ x.strftime('%d/%m/%Y') for x in date_list ]
                    dlist = list(set(dlist))
                    txt += f"List of attendance date matching to stages table:\n"
                    txt += str(dlist) + "\n"
                    df['timetable_date'] = df.apply(lambda x: str(x['timetable_date'])[:10],axis=1)
                    if sid==0:
                        cols = ['class_type_code','status_desc','attendance_type_desc','timetable_date','timetable_day','email']
                    else:
                        cols = get_columns("stages")
                        stage_list = sms_missingdates(client_name,course_id,sid,cols,date_list)
                        results = [x for x in stage_list if x != '']
                        txt += "Missing stages : " + str(results) + "\n"
                        att_rate = sms_att_rate(client_name,course_id,sid,date_list)
                        txt += f"Attendance rate = {att_rate}"
                        cols = ['class_type_code','status_desc','attendance_type_desc','timetable_date','timetable_day']
                    df1 = df[cols]
                    cmd_log += str(txt) + '\n'
                    cmd_log += str(str(df1)) + '\n'
    elif cmd=="xls_import" and resp != "":
        fname=resp
        cmd_log += f"Performaning course import for {fname}\n"
        (status,wb,master) = valid_format(fname,fname)
        if status==True:
            (status,msg,df) = course_import(client_name,fname,wb,master)
            cmd_log += str(msg) + '\n'
            if status:
                cmd_log += str(df) + '\n'
        else:
            cmd_log += str(wb) + '\n'
    elif cmd=="userinfo" and resp != "":
        if not resp.isnumeric():
            cmd_log += "Sorry it seems you are not a valid user in our system\n"
            return cmd_log
        qry = f"select username from user_master where client_name='{client_name}' and studentid={resp};"
        username = rds_param(qry)
        qry = "SELECT a.courseid FROM userdata a INNER JOIN playbooks b ON a.client_name=b.client_name "
        qry += f"AND a.courseid=b.course_id WHERE b.eoc=0 and a.studentid={resp} AND a.client_name = '{client_name}' LIMIT 1;"
        courseid = rds_param(qry)
        cmd_log += 'userinfo\n' + str(username) + '\n' + str(courseid) + '\n'
    elif cmd=="whois" and resp != "":
        qry = f"select * from user_master where client_name='{client_name}' and lower(username) like '%{resp}%';"
        df = rds_df(qry)
        df.columns = vmsvclib.get_columns("user_master")
        cmd_log += str(df) + '\n'
    elif cmd=="progress" and resp != "":
        if resp.isnumeric():
            qry = f"select studentid from user_master where client_name='{client_name}' and studentid={resp};"
        else:
            cmd_log += "Sorry it seems you are not a valid user in our system\n"
            return cmd_log
        studentid = rds_param(qry)
        sid = int("0" + str(studentid))
        if sid==0:
            cmd_log += str("Sorry it seems you are not a valid user in our system.") + '\n'
            return cmd_log
        else:
            vmbotlib.bot_intance = type('config',(object,), \
            {'pass_rate': 0.7,'att_rate':0.75, 'max_iu':20, 'gmt': 8})
            resp_dict = vmbotlib.load_respdict()
            qry = "select a.* FROM userdata a INNER JOIN playbooks b ON a.client_name=b.client_name AND a.courseid=b.course_id"
            qry += f" where b.eoc=0 and a.client_name='{client_name}' and a.studentid={resp};"
            userdata = rds_df(qry)
            if (userdata is None) or (len(userdata)==0):
                cmd_log += str("Unable to find the record in the system") + '\n'
                return cmd_log
            cols = vmsvclib.get_columns("userdata")
            userdata.columns = cols
            df = userdata.T
            df.columns = ['col']
            records = dict(zip(cols,[x for x in df.col]))
            course_id = records['courseid']
            qry = f"select * from stages where client_name='{client_name}' and courseid='{course_id}';"
            stagetable = rds_df(qry)
            if (stagetable is None) or (len(stagetable)==0):
                cmd_log += str("Unable to find the record in the system") + '\n'
                return cmd_log
            stagetable.columns = vmsvclib.get_columns("stages")
            vars = vmbotlib.display_progress(userdata, stagetable, sid, records, client_name, resp_dict, vmbotlib.bot_intance.pass_rate, vmbotlib.bot_intance.att_rate)
            if 'notification' in list(vars):
                cmd_log += str(vars['notification']) + '\n'
    elif cmd=="pipeline":
        generate_mcq_as(client_name)
        cmd_log += str("pipeline data generated") + '\n'
    elif cmd=="mass_update":
        load_edx(client_name)
        cmd_log += str("mass update completed") + '\n'
    else:
        cmd = ""
    if cmd == "":
        prog = "vmedxlib.py"
        cmd_log += str(f"usage :\n\tpython3 {prog} [commands] [cohort_id] [student_id]") + '\n'
        cmd_log += str("commands:\n\timport\n\tmcq\n\tassignment\n\tschedule\n\tcalendar\n\tinfo\n\tpipeline\n\tmass_update\n\txls_import\n\tprogress\n\tstages\n\tuserinfo\n\twhois\n") + '\n'
        cmd_log += str(f"Example:\n\tpython3 {prog} assignment IMM-0520A 4558") + '\n'
    return cmd_log

if __name__ == "__main__":
    global use_edxapi,edx_api_header,edx_api_url,max_iu,loop,cmd_log
    with open("vmbot.json") as json_file:
        bot_info = json.load(json_file)
    client_name = bot_info['client_name']
    edx_api_url = "https://omnimentor.sambaash.com/edx/v1"
    edx_api_header = {'Authorization': 'Basic ZWR4YXBpOlVzM3VhRUxJVXZENUU4azNXdG9E','Content-Type': 'text/plain'}
    vmsvclib.rds_connstr = bot_info['omdb']
    vmsvclib.rdsdb = None
    vmsvclib.rdscon = None
    vmsvclib.rds_pool = 0
    vmsvclib.rds_schema = bot_info['schema']
    max_iu = 20
    cmd_log = ""
    cmd = str(sys.argv[1]) if len(sys.argv)>=2 else ""
    resp = str(sys.argv[2]) if len(sys.argv)>=3 else ""
    arg = str(sys.argv[3]) if len(sys.argv)>=4 else ""
    cmd_log = parse_commands(client_name,cmd,resp,arg)
    print(cmd_log)
